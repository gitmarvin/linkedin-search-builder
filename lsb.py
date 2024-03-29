import openpyxl
import sys

workbook_name = sys.argv[1]
wb = openpyxl.load_workbook(workbook_name)
sh = wb[wb.sheetnames[0]]
prospect_list = []
checklist = []
reference_list = []


def get_prospect_list_single():
    name_column = input('Target column: ')
    for i in range(2, sh.max_row):
        prospect_name_single = sh.cell(row=i, column=int(name_column)).value
        prospect_name_formatted = '(' + '"' + prospect_name_single + '"' + ")"
        prospect_list.append(prospect_name_formatted)
        checklist.append(prospect_name_single)


def get_prospect_list_duple():
    column_first_name = input('Column with first name: ')
    column_last_name = input('Column with last name: ')
    for i in range(2, sh.max_row):
        prospect_first_name = sh.cell(row=i, column=int(column_first_name)).value
        prospect_last_name = sh.cell(row=i, column=int(column_last_name)).value
        #prospect_name_formatted = '(' + '"' + prospect_first_name + '"' + ' AND ' + '"' + prospect_last_name + '"' + ")"
        prospect_name_formatted = '(' + '"' + prospect_first_name  + ' ' + prospect_last_name + '"' + ")"
        #prospect_name_formatted = '(' + prospect_first_name + ' ' + prospect_last_name + ")"
        prospect_full_name = str(prospect_first_name) + ' ' + str(prospect_last_name)
        prospect_list.append(prospect_name_formatted)
        checklist.append(prospect_full_name)


def search_query_builder():
    search_string = ''
    for i in prospect_list:
        search_string += str(i) + ' OR '
    f = open('search-string.txt', "w")
    f.write(search_string)
    f.close()


def build_reference_list():
    reference_workbook = openpyxl.load_workbook('T1-Key-Accounts-06-19-reference.xlsx')
    reference_sheet = reference_workbook['Sheet 1 - T1-Key-Accounts-06-19']

    for i in range(3, 105):
        prospect_first_name = reference_sheet.cell(row=i, column=4).value
        prospect_last_name = reference_sheet.cell(row=i, column=5).value
        prospect_full_name = str(prospect_first_name) + ' ' + str(prospect_last_name)
        reference_list.append(prospect_full_name)


def check_missing_leads():
    for i in reference_list:
        if i in checklist and i is not None:
            checklist.remove(i)

    f = open('missing-leads.txt', 'w')
    for i in checklist:
        f.write(i + ' ')
        f.write('\n')
    f.close()


def dump_lists():
    f = open('list-dump.txt', 'w')

    f.write('Reference List' + '\n')
    f.write('\n')
    for i in reference_list:
        f.write(i)
        f.write('\n')
    f.write('\n')

    f.write('Checklist' + '\n')
    f.write('\n')
    for i in checklist:
        f.write(i)
        f.write('\n')
    f.write('\n')

    f.write('Prospect List' + '\n')
    f.write('\n')
    for i in prospect_list:
        f.write(i)
        f.write('\n')
    f.write('\n')

    f.close()


if __name__ == '__main__':

    type = input('Single or duple? (1/2): ')
    if type == '1':
        get_prospect_list_single()
        search_query_builder()
    elif type == '2':
        get_prospect_list_duple()
        search_query_builder()
    else:
        print('invalid input!')

    #print('Dumping lists..')
    #dump_lists()
    print('Operation completed!')
